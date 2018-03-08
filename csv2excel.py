#!/usr/bin/env python3

import os
import sys
import argparse
import csv
import re
from dateutil import parser
import datetime
from abc import ABC, abstractmethod

import xlwt
import openpyxl


class Csv2XlsException(Exception):
    pass


class ExcelWriter(ABC):

    @abstractmethod
    def add_sheet(self, wb, name):
        pass

    @abstractmethod
    def append(self, ws, row):
        pass

    @abstractmethod
    def save(self, filename):
        pass


class XlsWriter(ExcelWriter):

    def __init__(self):
        self.wb = xlwt.Workbook()
        self.row_counter = 0

    def add_sheet(self, name):
        self.row_counter = 0
        return self.wb.add_sheet(sheetname=name)

    def append(self, worksheet, row, translate_date=False):
        col_counter = 0
        for data in row:
            excel_data = parse_data(data, translate_date)
            if translate_date and type(excel_data) is datetime.datetime:
                datetime_format = xlwt.XFStyle()
                datetime_format.num_format_str = 'yyyy-MM-ddThh:mm:ss.000'
                worksheet.write(self.row_counter, col_counter, excel_data, datetime_format)
            elif translate_date and type(excel_data) is datetime.time:
                time_format = xlwt.XFStyle()
                time_format.num_format_str = 'hh:mm:ss.000'
                worksheet.write(self.row_counter, col_counter, excel_data, time_format)
            elif translate_date and type(excel_data) is datetime.date:
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'yyyy-MM-dd'
                worksheet.write(self.row_counter, col_counter, excel_data, date_format)
            else:
                worksheet.write(self.row_counter, col_counter, excel_data)
            col_counter += 1
        self.row_counter += 1

    def save(self, filename):
        self.wb.save(filename)


class XlsxWriter(ExcelWriter):

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.has_default_sheet = True

    def add_sheet(self, name):
        if self.has_default_sheet:
            self.has_default_sheet = False
            self.wb.active.title = name
            return self.wb.active
        return self.wb.create_sheet(title=name)

    def append(self, worksheet, row, translate_date=False):
        excel_row = [parse_data(data, translate_date) for data in row]
        worksheet.append(excel_row)

    def save(self, filename):
        self.wb.save(filename)


def parse_data(data, translate_date):
    try:
        return float(data)
    except ValueError:
        pass

    if translate_date:
        try:
            stripped_data = data.strip()
            dt = parser.parse(stripped_data, ignoretz=True)

            # look for time
            m = re.search('\d+:[0-5][0-9](:[0-5][0-9](.\d+)?)?$', stripped_data)
            if m:
                (time_start_idx, time_end_idx) = m.span()
                if time_start_idx == 0:
                    # if time start at index 0
                    # it means that there is only time on data
                    return datetime.time(dt.hour, dt.minute, dt.second, dt.microsecond)
                else:
                    # there is time and something else (maybe date)
                    return datetime.datetime(dt.year, dt.month, dt.day, dt.hour, dt.minute, dt.second, dt.microsecond)
            else:
                # no time found on date so only date
                return datetime.date(dt.year, dt.month, dt.day)
            return dt
        except ValueError:
            pass
    return data


def read_csv(csv_file):
    rows = []
    with open(csv_file, 'r', newline='') as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(row)
    return rows


def write_excel_from_csv(excel_file, csv_files, translate_date):
    dummy, file_extension = os.path.splitext(os.path.basename(excel_file))
    excel_writer = XlsWriter() if file_extension == '.xls' else XlsxWriter()
    for csv_file in csv_files:
        sheet_name = os.path.basename(csv_file)
        if sheet_name.lower().endswith('.csv'):
            sheet_name = sheet_name[:-4]
        rows = read_csv(csv_file)
        ws = excel_writer.add_sheet(sheet_name)
        for row in rows:
            excel_writer.append(ws, row, translate_date)
        excel_writer.save(excel_file)


def parse_arg():
    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--input-files',
                        nargs='+',
                        required=True,
                        help='input excel files')
    parser.add_argument('-o', '--output-file',
                        required=True,
                        help='output excel file')
    parser.add_argument('-t', '--translate-date',
                        action='store_true',
                        help='translate date string to excel date')

    args = parser.parse_args()
    return args


def main(argv):
    try:
        args = parse_arg()
        write_excel_from_csv(args.output_file, args.input_files, args.translate_date)
    except Exception as e:
        print(e, file=sys.stderr)
        exit(1)


if __name__ == "__main__":
    main(sys.argv)

